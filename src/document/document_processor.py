"""
문서 처리기
다양한 형식의 문서를 로드하고 청크로 분할
"""
import os
from pathlib import Path
from typing import List, Dict, Callable, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import (
    PyPDFLoader,
    Docx2txtLoader,
    UnstructuredWordDocumentLoader,
    TextLoader,
    CSVLoader,
    UnstructuredExcelLoader,
)

from .hwp_loader import HWPLoader, HWPXLoader

import sys
sys.path.append(str(Path(__file__).parent.parent.parent))
from config.settings import SUPPORTED_EXTENSIONS, CHUNK_SIZE, CHUNK_OVERLAP


class SmartTextLoader:
    """여러 인코딩을 시도하는 텍스트 로더"""
    
    ENCODINGS = ['utf-8', 'cp949', 'euc-kr', 'utf-16', 'latin-1']
    
    def __init__(self, file_path: str):
        self.file_path = file_path
    
    def load(self) -> List[Document]:
        """여러 인코딩으로 파일 로드 시도"""
        content = None
        used_encoding = None
        
        for encoding in self.ENCODINGS:
            try:
                with open(self.file_path, 'r', encoding=encoding) as f:
                    content = f.read()
                used_encoding = encoding
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        if content is None:
            # 바이너리로 읽어서 에러 무시
            try:
                with open(self.file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                used_encoding = 'utf-8 (with errors ignored)'
            except Exception as e:
                raise ValueError(f"파일을 읽을 수 없습니다: {e}")
        
        return [Document(
            page_content=content,
            metadata={"source": self.file_path, "encoding": used_encoding}
        )]


class SafeExcelLoader:
    """Excel 파일 안전하게 로드 (암호화된 파일도 지원)"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
    
    def load(self) -> List[Document]:
        """Excel 파일 읽기 (.xlsx는 openpyxl, .xls는 xlrd 또는 pandas)"""
        ext = Path(self.file_path).suffix.lower()
        
        if ext == '.xls':
            return self._load_xls()
        else:
            return self._load_xlsx()
    
    def _decrypt_if_needed(self, file_path: str):
        """암호화된 Excel 파일 복호화 (빈 비밀번호로 시도)"""
        try:
            import msoffcrypto
            import io
            
            with open(file_path, 'rb') as f:
                file_data = f.read()
            
            # 암호화 여부 확인
            file_obj = io.BytesIO(file_data)
            try:
                office_file = msoffcrypto.OfficeFile(file_obj)
                if office_file.is_encrypted():
                    # 빈 비밀번호로 복호화 시도 (Excel 기본 암호화)
                    decrypted = io.BytesIO()
                    try:
                        office_file.load_key(password='')
                        office_file.decrypt(decrypted)
                        return decrypted
                    except Exception:
                        # 비밀번호가 있는 파일
                        raise ValueError("비밀번호로 보호된 Excel 파일입니다. 암호를 해제 후 다시 시도하세요.")
            except Exception as e:
                if 'password' in str(e).lower() or '비밀번호' in str(e):
                    raise
                # 암호화되지 않은 파일 - 원본 반환
                pass
            
            return file_path
        except ImportError:
            # msoffcrypto 없으면 원본 파일 그대로 반환
            return file_path
    
    def _load_xlsx(self) -> List[Document]:
        """openpyxl로 .xlsx 읽기"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl이 설치되지 않았습니다. pip install openpyxl")
        
        # 암호화 파일 처리 시도
        file_to_load = self._decrypt_if_needed(self.file_path)
        
        try:
            wb = openpyxl.load_workbook(file_to_load, read_only=True, data_only=True)
        except Exception as e:
            error_msg = str(e).lower()
            if 'encrypted' in error_msg or 'password' in error_msg:
                raise ValueError(f"암호화된 Excel 파일입니다. 암호를 해제 후 다시 시도하세요.")
            raise ValueError(f"Excel 파일을 열 수 없습니다: {e}")
        
        documents = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                if row_text.strip():
                    rows.append(row_text)
            
            if rows:
                content = f"[시트: {sheet_name}]\n" + '\n'.join(rows)
                documents.append(Document(
                    page_content=content,
                    metadata={"source": self.file_path, "sheet": sheet_name}
                ))
        
        wb.close()
        return documents
    
    def _load_xls(self) -> List[Document]:
        """pandas로 .xls 읽기 (xlrd 필요)"""
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("pandas가 설치되지 않았습니다. pip install pandas")
        
        try:
            # xlrd 엔진으로 .xls 읽기
            xls = pd.ExcelFile(self.file_path, engine='xlrd')
        except ImportError:
            # xlrd 없으면 openpyxl 시도 (보통 실패하지만)
            try:
                xls = pd.ExcelFile(self.file_path)
            except Exception as e:
                raise ImportError(f"xlrd가 필요합니다. pip install xlrd: {e}")
        except Exception as e:
            raise ValueError(f"Excel 파일을 열 수 없습니다: {e}")
        
        documents = []
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            # DataFrame을 텍스트로 변환
            content = f"[시트: {sheet_name}]\n" + df.to_string(index=False)
            documents.append(Document(
                page_content=content,
                metadata={"source": self.file_path, "sheet": sheet_name}
            ))
        
        return documents


class DocumentProcessor:
    """문서 처리 총괄 클래스"""
    
    # 확장자별 로더 매핑
    LOADER_MAPPING = {
        '.hwp': HWPLoader,
        '.hwpx': HWPXLoader,
        '.pdf': PyPDFLoader,
        '.docx': Docx2txtLoader,
        '.doc': UnstructuredWordDocumentLoader,
        '.txt': SmartTextLoader,  # 여러 인코딩 시도
        '.md': SmartTextLoader,   # 여러 인코딩 시도
        '.csv': CSVLoader,
        '.xlsx': SafeExcelLoader,  # msoffcrypto 없이 동작
        '.xls': SafeExcelLoader,   # 구버전 Excel도 지원
    }
    
    def __init__(self, chunk_size: int = CHUNK_SIZE, chunk_overlap: int = CHUNK_OVERLAP):
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separators=["\n\n", "\n", "。", ".", " ", ""]
        )
    
    def scan_folder(self, folder_path: str) -> List[Path]:
        """폴더 내 지원되는 모든 문서 파일 스캔"""
        folder = Path(folder_path)
        if not folder.exists():
            raise FileNotFoundError(f"폴더를 찾을 수 없습니다: {folder_path}")
        
        files = []
        for ext in SUPPORTED_EXTENSIONS:
            files.extend(folder.rglob(f"*{ext}"))
        
        return sorted(files)
    
    def load_document(self, file_path: Path) -> List[Document]:
        """단일 문서 로드"""
        ext = file_path.suffix.lower()
        
        if ext not in self.LOADER_MAPPING:
            print(f"지원하지 않는 파일 형식: {file_path}")
            return []
        
        loader_class = self.LOADER_MAPPING[ext]
        
        try:
            loader = loader_class(str(file_path))
            documents = loader.load()
            
            # 메타데이터 보강
            for doc in documents:
                doc.metadata.update({
                    "source": str(file_path),
                    "filename": file_path.name,
                    "file_type": ext[1:],  # 확장자에서 점 제거
                })
            
            return documents
        except Exception as e:
            print(f"문서 로드 실패 ({file_path}): {e}")
            return []
    
    def process_folder(
        self, 
        folder_path: str, 
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> List[Document]:
        """
        폴더 내 모든 문서 처리
        
        Args:
            folder_path: 문서 폴더 경로
            progress_callback: 진행률 콜백 함수 (현재, 전체, 파일명)
        
        Returns:
            청크로 분할된 문서 목록
        """
        files = self.scan_folder(folder_path)
        total = len(files)
        
        if total == 0:
            return []
        
        all_documents = []
        
        for idx, file_path in enumerate(files, 1):
            if progress_callback:
                progress_callback(idx, total, file_path.name)
            
            docs = self.load_document(file_path)
            all_documents.extend(docs)
        
        # 청크로 분할
        if all_documents:
            chunked_docs = self.text_splitter.split_documents(all_documents)
            return chunked_docs
        
        return []
    
    def process_folder_parallel(
        self, 
        folder_path: str, 
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
        max_workers: int = 4
    ) -> List[Document]:
        """병렬 처리로 폴더 내 모든 문서 처리"""
        files = self.scan_folder(folder_path)
        total = len(files)
        
        if total == 0:
            return []
        
        all_documents = []
        completed = 0
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(self.load_document, f): f for f in files}
            
            for future in as_completed(futures):
                file_path = futures[future]
                completed += 1
                
                if progress_callback:
                    progress_callback(completed, total, file_path.name)
                
                try:
                    docs = future.result()
                    all_documents.extend(docs)
                except Exception as e:
                    print(f"처리 실패 ({file_path}): {e}")
        
        # 청크로 분할
        if all_documents:
            chunked_docs = self.text_splitter.split_documents(all_documents)
            return chunked_docs
        
        return []
    
    def get_supported_extensions(self) -> List[str]:
        """지원되는 파일 확장자 목록 반환"""
        return list(SUPPORTED_EXTENSIONS)
